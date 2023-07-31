#!/usr/bin/env python3

from datetime import date
import sys
import os
import argparse
import pathlib
import json
import shutil
import subprocess
import csv
import time
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

usersDict = {
    "apalladino": "Ali Palladino",
    "apecorale": "Aidan Pecorale",
    "azheng": "Alice Zheng",
    "bbruccoleri": "Bob Bruccoleri",
    "dprzybylski": "Darius Przybylski",
    "dselby": "Don Selby",
    "jprout": "Jamie Prout",
    "jvonstetina": "Jessica von Stetina",
    "JWright": "Jason Wright",
    "kgall": "Katie Gall",
    "mrubin": "Miranda Rubin",
    "nbourgeois": "Nate Bourgeois",
    "phaddadi": "Parham Haddadi",
    "spotts": "Sarah Potts",
    "thanscom": "Terrence Hanscom",
    "TStMartin": "Thia StMartin",
    "jira_api": "Aidan Pecorale"
}

def is_valid_file(parser, arg):
    if pathlib.Path(arg).suffix not in ('.xlsx', '.xls', '.json'):
        parser.error("Please upload a .xlsx or .json file")
    if not os.path.exists(arg):
        parser.error("The file %s does not exist." % arg)
    else:
        return arg  # return an open file handle

def get_arguments():
    parser = MyParser(description='Convert between Excel and JSON Nanopore basecalling samplesheet formats.',
                      formatter_class=MyHelpFormatter, add_help=False)

    required = parser.add_argument_group('Required')
    required.add_argument('-i', '--in_file', type=lambda x: is_valid_file(parser, x), required=True,
                          help='Input file')
    required.add_argument('-o', '--out_dir', type=pathlib.Path, required=True,
                          help='Output file path')

    options = parser.add_argument_group('Options')

    options.add_argument('--excel', action='store_true',
                         help='Convert JSON to Excel format (Default convert to JSON)')
    
    options.add_argument('--pipe', action='store_true',
                         help='Make csv for pipeline wrapper')

    options.add_argument('-h', '--help', action='help',
                         help='Show this help message and exit')

    args = parser.parse_args()
    check_arguments(args)
    return args


def main():
    # check_python_version()
    args = get_arguments()
    
    if args.excel:
        convert_to_excel(args)
        return

    resultJSON = convert_to_json(args)
    with open(args.out_file, "w") as outfile:
        outfile.write(resultJSON)

    return
    # if args.out_dir is not None:
        # make_output_directory(args.out_dir)


def convert_to_json(args):
    wb = load_workbook(args.in_file)
    ws = wb.active

    active_dict = {}

    expName = ws['B8'].value
    version = ws['B5'].value
    mode = ws['B6'].value
    flowcell = ws['B7'].value
    filterReads = ws['B9'].value
    filterReadsMinLen = ws['B10'].value
    filterReadsMinQual = ws['B11'].value
    date = ws['B12'].value
    experimentalistNano = list(usersDict.keys())[list(usersDict.values()).index(ws['B13'].value)] # untested

    active_dict["expName"] = expName
    active_dict["date"] = str(date).split()[0]
    active_dict["basecaller"] = version
    active_dict["basecallerMode"] = mode
    active_dict["flowcellID"] = flowcell
    active_dict["experimentalistNano"] = experimentalistNano

    if filterReads:
        active_dict["filterReadsMin"] = filterReadsMinLen
        active_dict["filterReadsMinQual"] = filterReadsMinQual

    projects_dict = {}
    rowCount = 0
    for row in ws.iter_rows(min_row=21, min_col=2, max_row=25, max_col=6):
        rowCount = rowCount + 1
        projName = 'project' + str(rowCount)
        tempProj = {}
        for id, cell in enumerate(row):
            if cell.value == '<Select>':
                break
            
            if id == 0:
                tempProj["hmiProj"] = cell.value

            if id == 1:
                tempProj["program"] = cell.value

            if id == 3:
                tempProj["briefDes"] = cell.value

            if id == 4:
                tempProj["des"] = cell.value


        if bool(tempProj):
            projects_dict[projName] = tempProj

            
    active_dict["projects"] = projects_dict

    samples_list = []
    rowCount = 0
    tempMultiplex = ""
    for row in ws.iter_rows(min_row=31, min_col=1, max_row=42, max_col=6):
        tempSample = {}
        if rowCount == 0:
            tempMultiplex = row[0].value
        rowCount = rowCount + 1
        for id, cell in enumerate(row):
            if id == 1:
                tempSample["barcodeNum"] = cell.value
            if id == 2:
                if cell.value is None:
                    break
                tempSample["id"] = cell.value

            if id == 3:
                tempSample["projNum"] = cell.value

            if id == 4:
                tempSample["constructId"] = cell.value

            if id == 5:
                tempSample["capsid"] = cell.value


        if bool(tempSample):
            tempSample["mplxKit"] = tempMultiplex
            samples_list.append(tempSample)

    rowCount = 0
    tempMultiplex = ""
    for row in ws.iter_rows(min_row=43, min_col=1, max_row=54, max_col=6):
        tempSample = {}
        if rowCount == 0:
            tempMultiplex = row[0].value
        rowCount = rowCount + 1
        if tempMultiplex == 'None':
            break
        for id, cell in enumerate(row):
            if id == 1:
                tempSample["barcodeNum"] = cell.value
            if id == 2:
                if cell.value is None:
                    break
                tempSample["id"] = cell.value

            if id == 3:
                tempSample["projNum"] = cell.value

            if id == 4:
                tempSample["constructId"] = cell.value

            if id == 5:
                tempSample["capsid"] = cell.value


        if bool(tempSample):
            tempSample["mplxKit"] = tempMultiplex
            samples_list.append(tempSample)
            
    active_dict["samples"] = samples_list


    return json.dumps(active_dict, indent=4)

def convert_to_excel(args):
    wb = load_workbook('pyScripts/samplesheet_template.xlsx')
    ws = wb['Form']

    with open(args.in_file) as json_file:
        data = json.load(json_file)

    ws['B5'] = data['basecaller']
    ws['B6'] = data['basecallerMode']
    ws['B7'] = data['flowcellID']
    ws['B8'] = data['expName']

    if 'filterReadsMin' not in data:
        ws['B9'] = 'False'
    else:
        ws['B10'] = data['filterReadsMin']
        ws['B11'] = data['filterReadsMinQual']

    # tempDate = 45024 # April 8th 2023

    # current date is number of days sinc Jan 0 1900
    f_date = date(1899, 12, 30) # adjusted to add missing day that is day of exp
    l_date = date(int(data['date'].split('-')[0]), int(data['date'].split('-')[1]), int(data['date'].split('-')[2]))

    delta = l_date - f_date

    ws['B12'] = delta.days
    # ws['B12'] = '45027'
    # ws['B12'] = data['date']
    # ws['B12'].number_format = 'mm/dd/yyyy'
   
    # other attempts
    # ws['B12'] = '45026'
    # date_style = NamedStyle(name='Date', number_format='MM/DD/YYYY')
    # ws['B12'].style = 'Date'
    
    modifiedExpName = usersDict[data['experimentalistNano']]

    ws['B13'] = modifiedExpName

    for projId, project in enumerate(data["projects"].keys()):
        row = 21 + projId
        ws['B' + str(row)] = data['projects'][project]['hmiProj']
        ws['C' + str(row)] = data['projects'][project]['program']
        ws['D' + str(row)] = modifiedExpName
        ws['E' + str(row)] = data['projects'][project]['briefDes']
        ws['F' + str(row)] = data['projects'][project]['des']

        ws['G' + str(row)] = 'Project_'+ data['projects'][project]['hmiProj'] + '_' + "".join(data['projects'][project]['program'].split()) + '_' + "".join(data['projects'][project]['briefDes'].split())


    for sampInd, sample in enumerate(data['samples']):
        row = 31 + int(sample['barcodeNum'][-2:]) - 1
        ws['C' + str(row)] = sample['id']
        ws['D' + str(row)] = sample['projNum']
        # if sample['constructId'] != '':
        #     ws['E' + str(row)] = sample['constructId']
        # if sample['capsid'] != '':
        #     ws['F' + str(row)] = sample['capsid']
        if 31 <= row <= 42:
            ws['A31'] = sample['mplxKit']
        if 43 <= row <= 54:
            ws['A43'] = sample['mplxKit']
    

    wb.save('{}.xlsx'.format(args.out_dir))

    if args.pipe:
        pipeline_prep(data, args.out_dir)

    return

def pipeline_prep(json_data, file_path):
    samples = json_data["samples"]
    projects = json_data["projects"]

    with open(file_path + ".csv", "w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["barcodeNum", "sampleName", "project"])

        for sample in samples:
            sample_id = sample["id"]
            barcode_num = sample["barcodeNum"][len(sample["barcodeNum"]) - 2:]
            proj_num = sample["projNum"].replace(" ", "")
            project = projects[proj_num]

            hmi_proj = project["hmiProj"]
            program = project["program"]
            brief_des = project["briefDes"]

            project_string = f"{hmi_proj}_{program}_{brief_des}"

            writer.writerow([barcode_num, sample_id, project_string.replace(" ", "")])

    print("CSV file generated successfully.")

def check_arguments(args):

    if args.out_dir is not None:
        if args.out_dir.is_file():
            sys.exit('Error: {} is a file (must be a directory)'.format(args.out_dir))

def get_colours_from_tput():
    try:
        return int(subprocess.check_output(['tput', 'colors']).decode().strip())
    except (ValueError, subprocess.CalledProcessError, FileNotFoundError, AttributeError):
        return 1

END_FORMATTING = '\033[0m'
BOLD = '\033[1m'
DIM = '\033[2m'

class MyParser(argparse.ArgumentParser):
    """
    This subclass of ArgumentParser changes the error messages, such that if a command is run with
    no other arguments, it will display the help text. If there is a different error, it will give
    the normal response (usage and error).
    """

    def error(self, message):
        if len(sys.argv) == 1:
            self.print_help(file=sys.stderr)
            sys.exit(2)
        else:
            super().error(message)


class MyHelpFormatter(argparse.HelpFormatter):
    """
    This is a custom formatter class for argparse. It adds some custom formatting like dim and bold.
    """

    def __init__(self, prog):
        terminal_width = shutil.get_terminal_size().columns
        os.environ['COLUMNS'] = str(terminal_width)
        max_help_position = min(max(24, terminal_width // 3), 40)
        self.colours = get_colours_from_tput()
        super().__init__(prog, max_help_position=max_help_position)

    def _get_help_string(self, action):
        """
        Override this function to add default values, but only when 'default' is not already in the
        help text.
        """
        help_text = action.help
        if action.default != argparse.SUPPRESS and action.default is not None:
            if 'default' not in help_text.lower():
                help_text += ' (default: {})'.format(action.default)
            elif 'default: DEFAULT' in help_text:
                help_text = help_text.replace('default: DEFAULT',
                                              'default: {}'.format(action.default))
        return help_text

    def start_section(self, heading):
        """
        Override this method to make section headers bold.
        """
        if self.colours > 1:
            heading = BOLD + heading + END_FORMATTING
        super().start_section(heading)

    def _format_action(self, action):
        """
        Override this method to make help descriptions dim.
        """
        help_position = min(self._action_max_length +
                            2, self._max_help_position)
        help_width = self._width - help_position
        action_width = help_position - self._current_indent - 2
        action_header = self._format_action_invocation(action)
        if not action.help:
            tup = self._current_indent, '', action_header
            action_header = '%*s%s\n' % tup
            indent_first = 0
        elif len(action_header) <= action_width:
            tup = self._current_indent, '', action_width, action_header
            action_header = '%*s%-*s  ' % tup
            indent_first = 0
        else:
            tup = self._current_indent, '', action_header
            action_header = '%*s%s\n' % tup
            indent_first = help_position
        parts = [action_header]
        if action.help:
            help_text = self._expand_help(action)
            help_lines = self._split_lines(help_text, help_width)
            first_line = help_lines[0]
            if self.colours > 8:
                first_line = DIM + first_line + END_FORMATTING
            parts.append('%*s%s\n' % (indent_first, '', first_line))
            for line in help_lines[1:]:
                if self.colours > 8:
                    line = DIM + line + END_FORMATTING
                parts.append('%*s%s\n' % (help_position, '', line))
        elif not action_header.endswith('\n'):
            parts.append('\n')
        for subaction in self._iter_indented_subactions(action):
            parts.append(self._format_action(subaction))
        return self._join_parts(parts)


if __name__ == '__main__':
    main()
